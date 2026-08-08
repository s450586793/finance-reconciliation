import importlib.util
from io import BytesIO
from pathlib import Path
from xml.etree import ElementTree
from zipfile import ZipFile

from openpyxl import load_workbook
from pypdf import PdfReader

GENERATOR_PATH = Path(__file__).with_name("generate_synthetic_fixtures.py")
GENERATED_SUFFIXES = {".xls", ".xlsx", ".pdf"}


def _load_generator():
    spec = importlib.util.spec_from_file_location(
        "synthetic_fixture_generator", GENERATOR_PATH
    )
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def _generated_files(root: Path) -> dict[Path, bytes]:
    return {
        path.relative_to(root): path.read_bytes()
        for path in root.rglob("*")
        if path.suffix in GENERATED_SUFFIXES
    }


def test_synthetic_fixture_generation_is_byte_stable_and_normalizes_xlsx(tmp_path):
    generator = _load_generator()
    output_dir = tmp_path / "fixtures"

    generator.generate_synthetic_fixtures(output_dir)
    first_generation = _generated_files(output_dir)
    generator.generate_synthetic_fixtures(output_dir)

    assert _generated_files(output_dir) == first_generation
    assert {path.suffix for path in first_generation} == GENERATED_SUFFIXES

    workbook_path = output_dir / "tax_input_invoices.xlsx"
    with ZipFile(workbook_path) as archive:
        assert all(
            entry.date_time == generator.FIXTURE_ZIP_TIMESTAMP
            for entry in archive.infolist()
        )
        core_properties = archive.read("docProps/core.xml")
    core = ElementTree.fromstring(core_properties)
    assert core.find("{http://purl.org/dc/elements/1.1/}creator").text == (
        "Finance Reconciliation synthetic fixture generator"
    )
    assert core.find("{http://purl.org/dc/terms/}modified").text == (
        "2026-01-01T00:00:00Z"
    )

    workbook = load_workbook(workbook_path, read_only=True)
    try:
        assert workbook.active.title == "发票基础信息"
        assert workbook.active.max_row == 3
    finally:
        workbook.close()
    assert len(PdfReader(BytesIO((output_dir / "invoice_00000000000000000001.pdf").read_bytes())).pages) == 1
